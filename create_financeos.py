from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import BarChart, Reference

OUTPUT_FILE = "FinanceOS_V1.xlsx"
MAX_ROWS = 2000
HEADER_FILL = PatternFill("solid", fgColor="404040")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


# ----------------------------- helpers -----------------------------

def style_header(ws, headers):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def apply_border_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def set_col_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def apply_number_format(ws, col_letter, start_row, end_row, number_format):
    for r in range(start_row, end_row + 1):
        ws[f"{col_letter}{r}"].number_format = number_format


def add_list_validation(ws, sqref, formula, allow_blank=True):
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    ws.add_data_validation(dv)
    dv.add(sqref)


def add_whole_number_validation(ws, sqref, min_value, max_value):
    dv = DataValidation(
        type="whole",
        operator="between",
        formula1=str(min_value),
        formula2=str(max_value),
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(sqref)


def add_decimal_validation(ws, sqref, min_value, max_value):
    dv = DataValidation(
        type="decimal",
        operator="between",
        formula1=str(min_value),
        formula2=str(max_value),
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(sqref)


def create_named_range(wb, name, ref):
    wb.defined_names.add(DefinedName(name=name, attr_text=ref))


def unique_preserve_order(values):
    seen = set()
    out = []
    for v in values:
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    return out


# ----------------------------- sheets -----------------------------

def create_sheet_90_categorias(wb):
    ws = wb.create_sheet("90_CATEGORIAS")
    headers = ["Categoria", "Subcategoria", "Grupo", "Ativo (Sim/Não)"]
    style_header(ws, headers)
    ws.freeze_panes = "A2"

    data = [
        ("Salário", "", "Receitas", "Sim"),
        ("Renda Extra", "", "Receitas", "Sim"),
        ("Moradia", "Aluguel", "Essenciais", "Sim"),
        ("Moradia", "Condomínio", "Essenciais", "Sim"),
        ("Alimentação", "Mercado", "Essenciais", "Sim"),
        ("Alimentação", "Restaurante", "Essenciais", "Sim"),
        ("Transporte", "Uber/Taxi", "Essenciais", "Sim"),
        ("Transporte", "Combustível", "Essenciais", "Sim"),
        ("Saúde", "Farmácia", "Essenciais", "Sim"),
        ("Saúde", "Plano de Saúde", "Essenciais", "Sim"),
        ("Investimentos", "Renda Fixa", "Dívidas/Poupar", "Sim"),
        ("Investimentos", "Ações/Fundos", "Dívidas/Poupar", "Sim"),
        ("Dívidas", "Empréstimos", "Dívidas/Poupar", "Sim"),
        ("Dívidas", "Cartão (juros)", "Dívidas/Poupar", "Sim"),
        ("Lazer", "Cinema/Shows", "Outros", "Sim"),
        ("Compras", "Roupas", "Outros", "Sim"),
        ("Educação", "Cursos", "Outros", "Sim"),
        ("Outros", "Diversos", "Outros", "Sim"),
    ]

    for i, row in enumerate(data, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    add_list_validation(ws, f"D2:D{MAX_ROWS}", "=NR_SIMNAO")

    set_col_widths(ws, {"A": 22, "B": 24, "C": 18, "D": 16})
    apply_border_range(ws, 1, len(data) + 1, 1, 4)

    categorias = unique_preserve_order([r[0] for r in data if r[3] == "Sim"])
    grupos = unique_preserve_order([r[2] for r in data])
    return categorias, grupos, data


def create_sheet_91_contas(wb):
    ws = wb.create_sheet("91_CONTAS")
    headers = ["Conta", "TipoConta", "SaldoInicial", "Ativo (Sim/Não)"]
    style_header(ws, headers)
    ws.freeze_panes = "A2"

    data = [
        ("Nubank", "Corrente", 0, "Sim"),
        ("Inter", "Corrente", 0, "Sim"),
        ("Dinheiro", "Dinheiro", 0, "Sim"),
        ("Investimentos", "Investimentos", 0, "Sim"),
    ]

    for i, row in enumerate(data, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    add_list_validation(ws, f"B2:B{MAX_ROWS}", '"Corrente,Poupança,Dinheiro,Investimentos"')
    add_list_validation(ws, f"D2:D{MAX_ROWS}", "=NR_SIMNAO")

    apply_number_format(ws, "C", 2, MAX_ROWS, 'R$ #,##0.00')
    set_col_widths(ws, {"A": 22, "B": 18, "C": 14, "D": 16})
    apply_border_range(ws, 1, len(data) + 1, 1, 4)

    contas = unique_preserve_order([r[0] for r in data if r[3] == "Sim"])
    return contas, data


def create_sheet_92_cartoes(wb):
    ws = wb.create_sheet("92_CARTOES")
    headers = ["Cartão", "FechamentoDia (1-31)", "VencimentoDia (1-31)", "Limite", "Ativo (Sim/Não)"]
    style_header(ws, headers)
    ws.freeze_panes = "A2"

    data = [("Nubank", 20, 25, 8000, "Sim")]

    for i, row in enumerate(data, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    add_whole_number_validation(ws, f"B2:C{MAX_ROWS}", 1, 31)
    add_list_validation(ws, f"E2:E{MAX_ROWS}", "=NR_SIMNAO")
    apply_number_format(ws, "D", 2, MAX_ROWS, 'R$ #,##0.00')

    set_col_widths(ws, {"A": 22, "B": 20, "C": 20, "D": 14, "E": 16})
    apply_border_range(ws, 1, len(data) + 1, 1, 5)

    cartoes = unique_preserve_order([r[0] for r in data if r[4] == "Sim"])
    return cartoes, data


def create_sheet_93_regras(wb):
    ws = wb.create_sheet("93_REGRAS")
    headers = ["ContémTexto", "Categoria", "Subcategoria", "Tipo (Receita/Despesa)", "Prioridade", "Ativo (Sim/Não)"]
    style_header(ws, headers)
    ws.freeze_panes = "A2"

    data = [
        ("UBER", "Transporte", "Uber/Taxi", "Despesa", 1, "Sim"),
        ("MERCADO", "Alimentação", "Mercado", "Despesa", 2, "Sim"),
        ("SALARIO", "Salário", "", "Receita", 1, "Sim"),
    ]

    for i, row in enumerate(data, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    add_list_validation(ws, f"B2:B{MAX_ROWS}", "=NR_CATEGORIAS")
    add_list_validation(ws, f"D2:D{MAX_ROWS}", "=NR_TIPO")
    add_list_validation(ws, f"F2:F{MAX_ROWS}", "=NR_SIMNAO")

    set_col_widths(ws, {"A": 24, "B": 20, "C": 24, "D": 24, "E": 12, "F": 16})
    apply_border_range(ws, 1, len(data) + 1, 1, 6)


def create_sheet_01_lancamentos(wb):
    ws = wb.create_sheet("01_LANCAMENTOS")
    headers = [
        "ID", "Data", "Descrição", "Tipo (Receita/Despesa)", "Categoria", "Subcategoria",
        "Conta", "FormaPgto", "Cartão", "Valor", "Status (Pago/Pendente)",
        "Observações", "Tags", "MêsRef", "AnoRef"
    ]
    style_header(ws, headers)
    ws.freeze_panes = "A2"

    samples = [
        [None, datetime(2026, 3, 5), "Salário mensal", "Receita", "Salário", "", "Nubank", "PIX", "", 6500, "Pago", "", "trabalho"],
        [None, datetime(2026, 3, 7), "Mercado do mês", "Despesa", "Alimentação", "Mercado", "Inter", "Débito", "", 850, "Pago", "Compras casa", "casa"],
        [None, datetime(2026, 3, 8), "Uber centro", "Despesa", "Transporte", "Uber/Taxi", "Nubank", "PIX", "", 42.5, "Pago", "", "mobilidade"],
        [None, datetime(2026, 3, 10), "Notebook parcelado", "Despesa", "Compras", "Roupas", "Nubank", "Crédito", "Nubank", 3200, "Pendente", "12x", "eletronico"],
        [None, datetime(2026, 3, 12), "Aporte investimentos", "Despesa", "Investimentos", "Renda Fixa", "Investimentos", "PIX", "", 1000, "Pago", "", "invest"],
    ]

    for i, row in enumerate(samples, start=2):
        ws.cell(row=i, column=2, value=row[1])
        ws.cell(row=i, column=3, value=row[2])
        ws.cell(row=i, column=4, value=row[3])
        ws.cell(row=i, column=5, value=row[4])
        ws.cell(row=i, column=6, value=row[5])
        ws.cell(row=i, column=7, value=row[6])
        ws.cell(row=i, column=8, value=row[7])
        ws.cell(row=i, column=9, value=row[8])
        ws.cell(row=i, column=10, value=row[9])
        ws.cell(row=i, column=11, value=row[10])
        ws.cell(row=i, column=12, value=row[11])
        ws.cell(row=i, column=13, value=row[12])

    for r in range(2, MAX_ROWS + 1):
        ws.cell(row=r, column=1, value=f"=ROW()-1")
        ws.cell(row=r, column=14, value=f'=TEXT(B{r},"yyyy-mm")')
        ws.cell(row=r, column=15, value=f"=YEAR(B{r})")

    add_list_validation(ws, f"D2:D{MAX_ROWS}", "=NR_TIPO")
    add_list_validation(ws, f"E2:E{MAX_ROWS}", "=NR_CATEGORIAS")
    add_list_validation(ws, f"G2:G{MAX_ROWS}", "=NR_CONTAS")
    add_list_validation(ws, f"H2:H{MAX_ROWS}", "=NR_FORMAPGTO")
    add_list_validation(ws, f"I2:I{MAX_ROWS}", "=NR_CARTOES")
    add_list_validation(ws, f"K2:K{MAX_ROWS}", "=NR_STATUS")

    apply_number_format(ws, "B", 2, MAX_ROWS, "dd/mm/yyyy")
    apply_number_format(ws, "J", 2, MAX_ROWS, 'R$ #,##0.00')

    set_col_widths(ws, {
        "A": 8, "B": 12, "C": 28, "D": 22, "E": 18, "F": 20, "G": 16,
        "H": 16, "I": 16, "J": 14, "K": 20, "L": 32, "M": 16, "N": 12, "O": 10,
    })
    apply_border_range(ws, 1, 30, 1, 15)


def create_sheet_02_cartao(wb):
    ws = wb.create_sheet("02_CARTAO")
    headers = [
        "DataCompra", "Descrição", "Categoria", "Subcategoria", "Cartão", "ValorTotal",
        "Parcelas", "PrimeiraFatura (yyyy-mm)", "Observações", "GeradoEmLancamentos (Sim/Não)"
    ]
    style_header(ws, headers)
    ws.freeze_panes = "A2"

    samples = [
        (datetime(2026, 3, 10), "Notebook parcelado", "Compras", "Roupas", "Nubank", 3200, 12, "2026-03", "Compra de trabalho", "Não"),
        (datetime(2026, 3, 15), "Curso online", "Educação", "Cursos", "Nubank", 600, 6, "2026-03", "", "Não"),
    ]

    for i, row in enumerate(samples, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    add_list_validation(ws, f"C2:C{MAX_ROWS}", "=NR_CATEGORIAS")
    add_list_validation(ws, f"D2:D{MAX_ROWS}", "=NR_SUBCATEGORIAS")
    add_list_validation(ws, f"E2:E{MAX_ROWS}", "=NR_CARTOES")
    add_whole_number_validation(ws, f"G2:G{MAX_ROWS}", 1, 36)
    add_list_validation(ws, f"J2:J{MAX_ROWS}", "=NR_SIMNAO")

    apply_number_format(ws, "A", 2, MAX_ROWS, "dd/mm/yyyy")
    apply_number_format(ws, "F", 2, MAX_ROWS, 'R$ #,##0.00')

    set_col_widths(ws, {"A": 12, "B": 28, "C": 18, "D": 22, "E": 16, "F": 14, "G": 10, "H": 18, "I": 32, "J": 24})
    apply_border_range(ws, 1, len(samples) + 1, 1, 10)


def create_sheet_04_orcamento(wb, categorias):
    ws = wb.create_sheet("04_ORCAMENTO")
    headers = ["Mês (yyyy-mm)", "Categoria", "Orçado", "Realizado", "Diferença", "Alerta%"]
    style_header(ws, headers)
    ws.freeze_panes = "A2"

    examples = [
        ("2026-03", categorias[0], 7000, 0, 0, 0.8),
        ("2026-03", "Moradia", 2500, 0, 0, 0.8),
        ("2026-03", "Alimentação", 1200, 0, 0, 0.8),
        ("2026-03", "Transporte", 500, 0, 0, 0.8),
        ("2026-03", "Saúde", 600, 0, 0, 0.8),
        ("2026-03", "Lazer", 400, 0, 0, 0.8),
    ]

    for i, row in enumerate(examples, start=2):
        ws.cell(row=i, column=1, value=row[0])
        ws.cell(row=i, column=2, value=row[1])
        ws.cell(row=i, column=3, value=row[2])
        ws.cell(row=i, column=4, value=f'=SUMIFS(\'01_LANCAMENTOS\'!$J:$J,\'01_LANCAMENTOS\'!$D:$D,"Despesa",\'01_LANCAMENTOS\'!$E:$E,B{i},\'01_LANCAMENTOS\'!$N:$N,A{i})')
        ws.cell(row=i, column=5, value=f"=C{i}-D{i}")
        ws.cell(row=i, column=6, value=row[5])

    for r in range(2, MAX_ROWS + 1):
        if ws.cell(row=r, column=4).value is None:
            ws.cell(row=r, column=4, value=f'=SUMIFS(\'01_LANCAMENTOS\'!$J:$J,\'01_LANCAMENTOS\'!$D:$D,"Despesa",\'01_LANCAMENTOS\'!$E:$E,B{r},\'01_LANCAMENTOS\'!$N:$N,A{r})')
            ws.cell(row=r, column=5, value=f"=C{r}-D{r}")

    add_list_validation(ws, f"A2:A{MAX_ROWS}", "=NR_MESES")
    add_list_validation(ws, f"B2:B{MAX_ROWS}", "=NR_CATEGORIAS")
    add_decimal_validation(ws, f"F2:F{MAX_ROWS}", 0, 1)

    apply_number_format(ws, "C", 2, MAX_ROWS, 'R$ #,##0.00')
    apply_number_format(ws, "D", 2, MAX_ROWS, 'R$ #,##0.00')
    apply_number_format(ws, "E", 2, MAX_ROWS, 'R$ #,##0.00')
    apply_number_format(ws, "F", 2, MAX_ROWS, "0%")

    set_col_widths(ws, {"A": 14, "B": 20, "C": 14, "D": 14, "E": 14, "F": 12})
    apply_border_range(ws, 1, len(examples) + 1, 1, 6)


def create_sheet_03_metas(wb):
    ws = wb.create_sheet("03_METAS")
    headers = ["Meta", "ValorAlvo", "ValorAtual", "Progresso", "DataLimite", "Prioridade", "Status", "Observações"]
    style_header(ws, headers)
    ws.freeze_panes = "A2"

    rows = [
        ("Reserva de Emergência", 20000, 4500, datetime(2027, 12, 31), "Alta", "Ativa", "Meta principal"),
        ("Entrada do Imóvel", 80000, 12000, datetime(2028, 6, 30), "Média", "Ativa", "Longo prazo"),
    ]

    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=row[0])
        ws.cell(row=i, column=2, value=row[1])
        ws.cell(row=i, column=3, value=row[2])
        ws.cell(row=i, column=4, value=f"=IFERROR(C{i}/B{i},0)")
        ws.cell(row=i, column=5, value=row[3])
        ws.cell(row=i, column=6, value=row[4])
        ws.cell(row=i, column=7, value=row[5])
        ws.cell(row=i, column=8, value=row[6])

    add_list_validation(ws, f"F2:F{MAX_ROWS}", '"Alta,Média,Baixa"')
    add_list_validation(ws, f"G2:G{MAX_ROWS}", '"Ativa,Concluída,Pausada"')

    apply_number_format(ws, "B", 2, MAX_ROWS, 'R$ #,##0.00')
    apply_number_format(ws, "C", 2, MAX_ROWS, 'R$ #,##0.00')
    apply_number_format(ws, "D", 2, MAX_ROWS, "0.00%")
    apply_number_format(ws, "E", 2, MAX_ROWS, "dd/mm/yyyy")

    set_col_widths(ws, {"A": 28, "B": 14, "C": 14, "D": 12, "E": 14, "F": 14, "G": 18, "H": 32})
    apply_border_range(ws, 1, len(rows) + 1, 1, 8)


def create_sheet_99_aux(wb, categorias, grupos, cat_data, contas, cartoes):
    ws = wb.create_sheet("99_AUX")
    ws["A1"] = "Lista"
    ws["B1"] = "Valor"
    ws["A1"].fill = HEADER_FILL
    ws["A1"].font = HEADER_FONT
    ws["B1"].fill = HEADER_FILL
    ws["B1"].font = HEADER_FONT

    # Sim/Não
    ws["A2"] = "NR_SIMNAO"
    ws["B2"] = "Sim"
    ws["B3"] = "Não"

    # Tipo
    ws["A5"] = "NR_TIPO"
    ws["B5"] = "Receita"
    ws["B6"] = "Despesa"

    # Status
    ws["A8"] = "NR_STATUS"
    ws["B8"] = "Pago"
    ws["B9"] = "Pendente"

    # Forma pagamento
    ws["A11"] = "NR_FORMAPGTO"
    formas = ["PIX", "Débito", "Crédito", "Dinheiro", "Boleto"]
    for i, v in enumerate(formas, start=11):
        ws.cell(row=i, column=2, value=v)

    # Meses
    ws["A18"] = "NR_MESES"
    month_row_start = 18
    month_values = []
    for year in [2026, 2027]:
        for month in range(1, 13):
            month_values.append(f"{year}-{month:02d}")
    for i, m in enumerate(month_values, start=month_row_start):
        ws.cell(row=i, column=2, value=m)

    # Listas únicas auxiliares
    ws["D1"] = "CategoriasAtivas"
    ws["E1"] = "Grupos"
    ws["F1"] = "ContasAtivas"
    ws["G1"] = "CartoesAtivos"
    ws["H1"] = "Subcategorias"

    for i, c in enumerate(categorias, start=2):
        ws.cell(row=i, column=4, value=c)
    for i, g in enumerate(grupos, start=2):
        ws.cell(row=i, column=5, value=g)
    for i, c in enumerate(contas, start=2):
        ws.cell(row=i, column=6, value=c)
    for i, c in enumerate(cartoes, start=2):
        ws.cell(row=i, column=7, value=c)

    subcategorias = [r[1] for r in cat_data if r[1]]
    for i, s in enumerate(subcategorias, start=2):
        ws.cell(row=i, column=8, value=s)

    # Dashboard tables
    ws["J1"] = "Tabelas para Dashboard"
    ws["J2"] = "Mês selecionado"
    ws["K2"] = "2026-03"
    ws["J3"] = "TotalReceitasMes"
    ws["K3"] = '=SUMIFS(\'01_LANCAMENTOS\'!$J:$J,\'01_LANCAMENTOS\'!$D:$D,"Receita",\'01_LANCAMENTOS\'!$N:$N,K2)'
    ws["J4"] = "TotalDespesasMes"
    ws["K4"] = '=SUMIFS(\'01_LANCAMENTOS\'!$J:$J,\'01_LANCAMENTOS\'!$D:$D,"Despesa",\'01_LANCAMENTOS\'!$N:$N,K2)'
    ws["J5"] = "SaldoMes"
    ws["K5"] = "=K3-K4"

    ws["J7"] = "Categoria"
    ws["K7"] = "Valor"
    ws["J7"].fill = HEADER_FILL
    ws["J7"].font = HEADER_FONT
    ws["K7"].fill = HEADER_FILL
    ws["K7"].font = HEADER_FONT

    for i, cat in enumerate(categorias, start=8):
        ws.cell(row=i, column=10, value=cat)
        ws.cell(
            row=i,
            column=11,
            value=f'=SUMIFS(\'01_LANCAMENTOS\'!$J:$J,\'01_LANCAMENTOS\'!$D:$D,"Despesa",\'01_LANCAMENTOS\'!$E:$E,J{i},\'01_LANCAMENTOS\'!$N:$N,K2)'
        )

    ws["J30"] = "Instruções"
    ws["J31"] = "No Google Sheets, Apps Script gera parcelas em 01_LANCAMENTOS."
    ws["J33"] = "Exemplo divisão"
    ws["J34"] = "ValorTotal"
    ws["K34"] = 1200
    ws["J35"] = "Parcelas"
    ws["K35"] = 6
    ws["J36"] = "ValorParcela"
    ws["K36"] = "=IFERROR(K34/K35,0)"

    add_list_validation(ws, "K2", "=NR_MESES")
    apply_number_format(ws, "K", 3, 5, 'R$ #,##0.00')
    apply_number_format(ws, "K", 8, 8 + len(categorias), 'R$ #,##0.00')
    apply_number_format(ws, "K", 34, 36, 'R$ #,##0.00')

    set_col_widths(ws, {
        "A": 16, "B": 18, "D": 20, "E": 18, "F": 18, "G": 18, "H": 20,
        "J": 28, "K": 18,
    })

    # Named ranges
    create_named_range(wb, "NR_SIMNAO", "'99_AUX'!$B$2:$B$3")
    create_named_range(wb, "NR_TIPO", "'99_AUX'!$B$5:$B$6")
    create_named_range(wb, "NR_STATUS", "'99_AUX'!$B$8:$B$9")
    create_named_range(wb, "NR_FORMAPGTO", "'99_AUX'!$B$11:$B$15")
    create_named_range(wb, "NR_MESES", "'99_AUX'!$B$18:$B$41")
    create_named_range(wb, "NR_CATEGORIAS", f"'99_AUX'!$D$2:$D${1 + len(categorias)}")
    create_named_range(wb, "NR_GRUPOS", f"'99_AUX'!$E$2:$E${1 + len(grupos)}")
    create_named_range(wb, "NR_CONTAS", f"'99_AUX'!$F$2:$F${1 + len(contas)}")
    create_named_range(wb, "NR_CARTOES", f"'99_AUX'!$G$2:$G${1 + len(cartoes)}")
    create_named_range(wb, "NR_SUBCATEGORIAS", f"'99_AUX'!$H$2:$H${1 + len(subcategorias)}")


def create_sheet_00_dashboard(wb, categorias):
    ws = wb.create_sheet("00_DASHBOARD", 0)

    ws["A1"] = "FinanceOS V1 - Dashboard"
    ws["A1"].font = Font(size=18, bold=True, color="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="left")

    ws["A2"] = "Mês de análise"
    ws["B2"] = "2026-03"
    add_list_validation(ws, "B2", "=NR_MESES")

    # conecta seleção ao AUX
    wb["99_AUX"]["K2"] = "='00_DASHBOARD'!B2"

    # Cards
    cards = [
        ("B4", "Receitas do mês", "='99_AUX'!K3"),
        ("E4", "Despesas do mês", "='99_AUX'!K4"),
        ("H4", "Saldo do mês", "='99_AUX'!K5"),
        ("K4", "% orçamento usado", '=IFERROR(\'99_AUX\'!K4/SUMIFS(\'04_ORCAMENTO\'!$C:$C,\'04_ORCAMENTO\'!$A:$A,B2),0)'),
    ]

    for start_cell, title, formula in cards:
        col = ws[start_cell].column
        row = ws[start_cell].row
        ws.cell(row=row, column=col, value=title)
        ws.cell(row=row + 1, column=col, value=formula)

        for rr in [row, row + 1]:
            c = ws.cell(row=rr, column=col)
            c.fill = PatternFill("solid", fgColor="D9E1F2")
            c.font = Font(bold=(rr == row), color="1F1F1F")
            c.alignment = Alignment(horizontal="center")
            c.border = THIN_BORDER

    ws["B5"].number_format = 'R$ #,##0.00'
    ws["E5"].number_format = 'R$ #,##0.00'
    ws["H5"].number_format = 'R$ #,##0.00'
    ws["K5"].number_format = "0.00%"

    # Área do gráfico
    ws["A8"] = "Despesas por categoria"
    ws["A8"].font = Font(size=13, bold=True)

    chart = BarChart()
    chart.title = "Despesas por categoria"
    chart.style = 10
    chart.y_axis.title = "Valor (R$)"
    chart.x_axis.title = "Categoria"

    data = Reference(wb["99_AUX"], min_col=11, min_row=7, max_row=7 + len(categorias))
    cats = Reference(wb["99_AUX"], min_col=10, min_row=8, max_row=7 + len(categorias))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    ws.add_chart(chart, "A9")

    set_col_widths(ws, {
        "A": 24, "B": 18, "C": 4, "D": 4, "E": 18, "F": 4, "G": 4,
        "H": 18, "I": 4, "J": 4, "K": 20,
    })


def reorder_sheets(wb):
    desired = [
        "00_DASHBOARD",
        "01_LANCAMENTOS",
        "02_CARTAO",
        "03_METAS",
        "04_ORCAMENTO",
        "90_CATEGORIAS",
        "91_CONTAS",
        "92_CARTOES",
        "93_REGRAS",
        "99_AUX",
    ]
    wb._sheets = [wb[name] for name in desired]


def main():
    wb = Workbook()
    wb.remove(wb.active)

    categorias, grupos, cat_data = create_sheet_90_categorias(wb)
    contas, _ = create_sheet_91_contas(wb)
    cartoes, _ = create_sheet_92_cartoes(wb)
    create_sheet_93_regras(wb)
    create_sheet_01_lancamentos(wb)
    create_sheet_02_cartao(wb)
    create_sheet_04_orcamento(wb, categorias)
    create_sheet_03_metas(wb)
    create_sheet_99_aux(wb, categorias, grupos, cat_data, contas, cartoes)
    create_sheet_00_dashboard(wb, categorias)

    reorder_sheets(wb)
    wb.save(OUTPUT_FILE)
    print(f"Arquivo gerado com sucesso: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
